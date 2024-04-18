import pandas as pd
from openpyxl import load_workbook

def getMatriz(archivo):
    wb = load_workbook(archivo)
    hoja = wb['Hoja1']
    filas = hoja.max_row
    columnas = hoja.max_column

    dicc_relaciones = {}
    for row in range(2, filas + 1):
        nodo = hoja.cell(row=row, column=1).value
        dicc_relaciones[nodo] = {}
        for col in range(2, columnas + 1):
            peso = hoja.cell(row=row, column=col).value
            relacion = hoja.cell(row=1, column=col).value
            if relacion is not None and peso != 0:
                dicc_relaciones[nodo][relacion] = peso

    return dicc_relaciones


class Grafo:
    def __init__(self, archivo) -> None:
        self.Aristas = self.Excel(archivo)
    
    def Excel(self, archivo):
        df = pd.read_excel(archivo, index_col=0)
        return df.to_dict(orient='index')
    
    def addVertice(self, vertice):
        if vertice not in self.Aristas:
            self.Aristas[vertice] = {}
    
    def addArista(self, origen, destino, peso):
        self.addVertice(origen)
        self.addVertice(destino)
        self.Aristas[origen][destino] = peso
    
    def __str__(self) -> str:
        return self.printDicc(self.Aristas)

    def printDicc(self, dicc):
        result = ""
        for i in dicc:
            result += f'nodo: {i}\n'
            for j in dicc[i]:
                result += f'\tDestino:{j},peso:{dicc[i][j]}\n'
        return result

#---------main-----------

print("\033[34m Xeli Vargas \033[0m")
archivo = input("¿Cuál es el nombre del archivo a procesar? ")


# archivo Excel
print("\033[34m Contenido del Archivo Excel :\033[0m")
wb = load_workbook(archivo)
for hoja in wb.sheetnames:
    hoja = wb[hoja]
    for row in hoja.iter_rows(values_only=True):
        print("\t".join(str(cell) for cell in row))

#  lista de relaciones
print("\n\033[34m lista relaciones :\033[0m")
listarel= getMatriz(archivo)
for nodo, relaciones in listarel.items():
    relaciones_str = " ".join([f"[{relacion}] : {peso} ->" for relacion, peso in relaciones.items()]) 
    print(f"Nodo {nodo} -> {relaciones_str}")


grafoTest = Grafo(archivo)
print("")

print("\n\033[34m GRAFO:\033[0m")
print(grafoTest)

origenG = 'B'
destinoG = 'H'
path = {}
visitados = []

visitados.append(origenG)
path[origenG] = {'-': 0}

verticeAct = origenG
while verticeAct != destinoG:
    visitados.append(verticeAct)
    llaves = grafoTest.Aristas[verticeAct].keys()

    for i in llaves:
        if i not in visitados:
            if i not in path:
                path[i] = {}
            
            # Calcula la nueva acumulación
            llave = list(path[verticeAct].keys())
            acumulado = path[verticeAct][llave[0]]
            if grafoTest.Aristas[verticeAct][i] != 0:
                nueva_acumulacion = grafoTest.Aristas[verticeAct][i] + acumulado
                if i not in path or nueva_acumulacion < path[i].get(verticeAct, float('inf')):
                    path[i][verticeAct] = nueva_acumulacion

            

    min_peso = float('inf')
    for nodo, pesos in path.items():
        if nodo not in visitados:
            for v, p in pesos.items():
                if p < min_peso:
                    min_peso = p
                    verticeAct = nodo

camino_mas_corto = [destinoG]
vertice_actual = destinoG
peso_total = 0  
while vertice_actual != origenG:
    for v, p in path[vertice_actual].items():
        if p == min_peso:
            camino_mas_corto.append(v)
            peso_total += grafoTest.Aristas[v][vertice_actual]  # Sumamos el peso de la arista al peso total
            min_peso -= grafoTest.Aristas[v][vertice_actual]
            vertice_actual = v
            break
camino_mas_corto.reverse()

# Imprimir el peso total del camino más corto
print(f"Peso total del camino más corto: {peso_total}")

# Imprimir el resultado
print("Camino más corto:")
print(camino_mas_corto)
"""print(f"Peso total del camino: {path[destinoG][origenG]}")"""
print("El peso Total es:",path)
print (visitados)
