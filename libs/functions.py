from openpyxl import load_workbook
import pandas as pd 

def getMatriz(archivo):
    wb = load_workbook(archivo)
    hoja = wb['Hoja1']
    filas = hoja.max_row
    columnas = hoja.max_column

    dicc_relaciones = {}
    for row in range(2, filas + 1):
        nodo = hoja.cell(row=row, column=1).value
        dicc_relaciones[nodo] = {}
        for col in range(2, columnas + 1):
            peso = hoja.cell(row=row, column=col).value
            relacion = hoja.cell(row=1, column=col).value
            if relacion is not None and peso != 0:
                dicc_relaciones[nodo][relacion] = peso

    return dicc_relaciones


class Grafo:
    def __init__(self, archivo) -> None:
        self.Aristas = self.Excel(archivo)
    
    def Excel(self, archivo):
        df = pd.read_excel(archivo, index_col=0)
        return df.to_dict(orient='index')
    
    def addVertice(self, vertice):
        if vertice not in self.Aristas:
            self.Aristas[vertice] = {}
    
    def addArista(self, origen, destino, peso):
        self.addVertice(origen)
        self.addVertice(destino)
        self.Aristas[origen][destino] = peso
    
    def __str__(self) -> str:
        return self.printDicc(self.Aristas)

    def printDicc(self, dicc):
        result = ""
        for i in dicc:
            result += f'nodo: {i}\n'
            for j in dicc[i]:
                result += f'\tDestino:{j},peso:{dicc[i][j]}\n'
        return result




